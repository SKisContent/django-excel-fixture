"""
Serializes a Django Model to Excel and vice-versa
Some observations:
OpenPyXL offers both reference addressing (e.g., 'A1', 'AC93') and indexed addressing.
When indexing into a row, the base (first) index is 1. However, after that you're
in the world of python tuples, so the indexes are 0-based (e.g., ws[1][0] == ws['A1'], ws[2][4] == ws['E2'])
The same way, columns can be directly referenced (ws['A']), but then rows are 0-based
(e.g., ws['A'][0] == ws['A1'], ws['C'][4] == ws['C5'])
"""
import decimal
import io
import re
from datetime import datetime, date

import pytz
from django.apps import apps
from django.db import models
from django.conf import settings
from django.core.serializers import base
from django.utils import timezone
from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import save_workbook


PREFERRED_TS_FORMAT = '%Y-%m-%dT%H:%M:%S:%f%z'
Y_M_D_FORMAT = '%Y-%m-%d'

DATETIME_FORMATS = {r'\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}.\d{6}[-+]\d{4}':PREFERRED_TS_FORMAT,
                    r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}.\d{6}':'%Y-%m-%d %H:%M:%S:%f',
                    r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}':'%Y-%m-%d %H:%M:%S',
                    }


class Serializer(base.Serializer):
    def start_serialization(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.current_row = 2

    def start_object(self, obj):
        if not hasattr(obj, "_meta"):
            raise base.SerializationError("Non-model object (%s) encountered during serialization" % type(obj))

        # The first time this method is called, write the model object and field names
        if self.current_row < 3:
            self.ws['A1'] = obj._meta.label
            self.model_fields = dict([(mf.name, mf) for mf in obj._meta.fields if mf.serialize])
            self.mf_keys = list(self.model_fields.keys())
            self.ws.append(self.mf_keys) # appends a new row with the values in the iterable
            self.field_positions = {key: index for index, key in enumerate(self.mf_keys)}
            self.num_fields = len(self.mf_keys)
        # prepare for another row of field values
        self.current_row += 1

    def handle_field(self, obj, field):
        """
        Do any special handling of field values. For example, save datetime values with timestamps
        :param obj:
        :param field:
        :return:
        """
        if getattr(obj, field.name) is not None:
            try:
                if isinstance(field, models.BooleanField):
                    value = field.value_from_object(obj)
                elif isinstance(field, models.DateTimeField):
                    value = field.value_from_object(obj).strftime(PREFERRED_TS_FORMAT)
                elif isinstance(field, models.DateField):
                    value = field.value_from_object(obj)
                elif isinstance(field, models.DecimalField):
                    value = field.value_from_object(obj)
                else:
                    # Get a "string version" of the object's data.
                    value = field.value_to_string(obj)
                self.ws[self.current_row][self.field_positions[field.name]].value = value
            except base.SerializationError:
                raise ValueError("%s.%s (pk:%s) contains unserializable characters" %
                                        (obj.__class__.__name__, field.name, obj.pk))
        else:
            # ToDo: do nothing or add a None-signifying value to the sheet?
            pass

    def handle_fk_field(self, obj, field):
        pass

    def handle_m2m_field(self, obj, field):
        pass

    def end_serialization(self):
        filename = self.stream.name
        if filename == '<stdout>':
            # If there is no file, dump a CSV representation to stdout
            out_str = self.csv(self.ws)
            self.stream.write(out_str)
        else:
            # The default stream is opened in text mode, but we need binary
            self.stream.close()
            # Just use the openpyxl saving method
            save_workbook(self.wb, filename)

    def csv(self, worksheet):
        """

        :param worksheet:
        :return:
        """
        num_rows = len(worksheet['A'])
        if num_rows == 1:
            raise base.SerializationError('There is no data to dump.')
        buffer = io.StringIO()
        for ix in range(num_rows):
            buffer.write(','.join([wrap(cell.value) if cell.value else '' for cell in worksheet[ix+1]]))
            buffer.write('\n')
        return buffer.getvalue()


def wrap(value):
    if value is None:
        return None
    if type(value) is str:
        value.replace('"','""')
    else:
        value = str(value)
    return '"{0}"'.format(value)


class Deserializer(base.Deserializer):
    def __init__(self, stream_or_string, **options):
        super(Deserializer, self).__init__(stream_or_string, **options)
        wb = load_workbook(stream_or_string)
        self.ws = wb.active
        self.model_identifier = self.ws['A1'].value
        self.Model = self._get_model(self.model_identifier)
        self.model_fields = dict([(mf.name, mf) for mf in self.Model._meta.fields])
        self.num_fields = len(self.ws['1'])
        self.num_objects = len(self.ws['A']) - 2
        self.fields = [(cell.value, self.model_fields[cell.value]) for cell in self.ws[2]]
        self.auto_now_fields = [af for af in self.Model._meta.fields
                                if (hasattr(af, 'auto_now') and af.auto_now) or
                                   (hasattr(af, 'auto_now_add') and af.auto_now_add)]

        self.current_row = 3 # First row is 1

    def __next__(self):
        if self.current_row < self.num_objects + 3:
            values = {}
            for ix in range(self.num_fields):
                values[self.fields[ix][0]] = self.get_value(self.ws[self.current_row][ix], self.fields[ix][1])
            present = datetime.now()
            # The following is not necessary since we are saving the object using it's save method
            # However, it doesn't hurt
            for af in self.auto_now_fields:
                if af.name not in values or values[af.name] is None:
                    values[af.name] = present
            obj = base.build_instance(self.Model, values, False)
            self.current_row += 1
            obj.save()
            return base.DeserializedObject(obj, {})
        raise StopIteration

    def _get_model(self, model_identifier):
        """
        Look up model
        """
        if not model_identifier:
            raise base.DeserializationError("Worksheet is missing the required model name")
        try:
            return apps.get_model(model_identifier)
        except (LookupError, TypeError):
            raise base.DeserializationError("Worksheet has invalid model identifier: '%s'"
                % (model_identifier))

    def get_value(self, cell, field):
        if cell.value is None:
            return None
        if str(cell.value).startswith('='):
            raise base.DeserializationError("Formulas are not supported at this time. Cell %s%s" % (cell.row, cell.column))
        if isinstance(field, models.BooleanField):
            return cell.value if type(cell.value) is bool else bool(cell.value)
        elif isinstance(field, models.DateTimeField):
            # Order matters, DateTime needs to come before Date because datetime subclasses date
            # isinstance(aDateTimeInstance, DateField) is True, isinstance(aDateInstance, DateTimeField) is False
            if type(cell.value) is datetime:
                if settings.USE_TZ and timezone.is_naive(cell.value):
                    return timezone.make_aware(cell.value, pytz.UTC)
                else:
                    return cell.value
            else:
                # Handle a couple different timestamp formats -- first one is how *we* save 'em
                for pattern, format in DATETIME_FORMATS.items():
                    if re.match(pattern, str(cell.value)):
                        return datetime.strptime(str(cell.value), format)
                return str(cell.value)  # hope for the best
        elif isinstance(field, models.DateField):
            if type(cell.value) is datetime:
                return cell.value.date()
            else:
                return cell.value if type(cell.value) is date else datetime.strptime(cell.value, Y_M_D_FORMAT).date()
        elif isinstance(field, models.DecimalField):
            try:
                return decimal.Decimal(cell.value)
            except Exception as ex:
                print(cell.value)
        else:
            return str(cell.value)
