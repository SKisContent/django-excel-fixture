# coding: utf-8
import datetime
from django.db import models
from io import BytesIO
from unittest import skip
from unittest.mock import patch
from django.core import serializers
from django.test import TestCase
from openpyxl import Workbook
from openpyxl.writer.excel import save_workbook
from tests.myapp.models import Person, Recipe, Ingredient, Association

class ExcelFileGenMixin:

    def _get_xlsx_stream(self):
        """ build a xlsx file in memory, based on the test app model."""
        workbook = Workbook()

        # Remove auto-created sheet:
        name = workbook.get_sheet_names()[0]
        sheet = workbook.get_sheet_by_name(name)
        workbook.remove_sheet(sheet)

        # Add sheet:
        workbook.create_sheet('myapp.Person')
        # Add header:
        workbook['myapp.Person'].cell(row=1, column=1, value='id')
        workbook['myapp.Person'].cell(row=1, column=2, value='name')
        workbook['myapp.Person'].cell(row=1, column=3, value='age')
        # Add content 1:
        workbook['myapp.Person'].cell(row=2, column=1, value=1)
        workbook['myapp.Person'].cell(row=2, column=2, value='Person 1')
        workbook['myapp.Person'].cell(row=2, column=3, value=21)
        # Add content 2:
        workbook['myapp.Person'].cell(row=3, column=1, value=2)
        workbook['myapp.Person'].cell(row=3, column=2, value='Person 2')
        workbook['myapp.Person'].cell(row=3, column=3, value=22)
        # Add content 3:
        workbook['myapp.Person'].cell(row=4, column=1, value=3)
        workbook['myapp.Person'].cell(row=4, column=2, value='Person 3')
        workbook['myapp.Person'].cell(row=4, column=3, value=23)

        # Add sheet:
        workbook.create_sheet('myapp.Ingredient')
        # Add header:
        workbook['myapp.Ingredient'].cell(row=1, column=1, value='id')
        workbook['myapp.Ingredient'].cell(row=1, column=2, value='name')
        # Add content 1:
        workbook['myapp.Ingredient'].cell(row=2, column=1, value=1)
        workbook['myapp.Ingredient'].cell(row=2, column=2, value='Ingredient 1')
        # Add content 2:
        workbook['myapp.Ingredient'].cell(row=3, column=1, value=2)
        workbook['myapp.Ingredient'].cell(row=3, column=2, value='Ingredient 2')
        # Add content 3:
        workbook['myapp.Ingredient'].cell(row=4, column=1, value=3)
        workbook['myapp.Ingredient'].cell(row=4, column=2, value='Ingredient 3')
        # Add content 4:
        workbook['myapp.Ingredient'].cell(row=5, column=1, value=4)
        workbook['myapp.Ingredient'].cell(row=5, column=2, value='Ingredient 4')

        # Add sheet:
        workbook.create_sheet('myapp.Recipe')
        # Add header:
        workbook['myapp.Recipe'].cell(row=1, column=1, value='id')
        workbook['myapp.Recipe'].cell(row=1, column=2, value='name')
        workbook['myapp.Recipe'].cell(row=1, column=3, value='owner')
        workbook['myapp.Recipe'].cell(row=1, column=4, value='cooking_time')
        workbook['myapp.Recipe'].cell(row=1, column=5, value='created_at')
        workbook['myapp.Recipe'].cell(row=1, column=6, value='rtype')
        ## Add content 1:
        workbook['myapp.Recipe'].cell(row=2, column=1, value=1)
        workbook['myapp.Recipe'].cell(row=2, column=2, value='Recipe 1')
        workbook['myapp.Recipe'].cell(row=2, column=3, value=1)
        workbook['myapp.Recipe'].cell(row=2, column=4, value=None)
        workbook['myapp.Recipe'].cell(row=2, column=5, value='2019-09-10 15:17:34:623000')
        workbook['myapp.Recipe'].cell(row=2, column=6, value='V')
        # Add content 2:
        workbook['myapp.Recipe'].cell(row=3, column=1, value=2)
        workbook['myapp.Recipe'].cell(row=3, column=2, value='Recipe 2')
        workbook['myapp.Recipe'].cell(row=3, column=3, value=1)
        workbook['myapp.Recipe'].cell(row=3, column=4, value=None)
        workbook['myapp.Recipe'].cell(row=3, column=5, value='2019-09-10 15:17:34:623000')
        workbook['myapp.Recipe'].cell(row=3, column=6, value='V')


        # Add sheet:
        workbook.create_sheet('myapp.Association')
        # Add header:
        workbook['myapp.Association'].cell(row=1, column=1, value='id')
        workbook['myapp.Association'].cell(row=1, column=2, value='recipe')
        workbook['myapp.Association'].cell(row=1, column=3, value='ingredient')
        workbook['myapp.Association'].cell(row=1, column=4, value='amount')


        # Generate in memory stream:
        in_memory_file = BytesIO()
        save_workbook(workbook, in_memory_file)

        return in_memory_file


class XlsxDeserializerFunctionalTest(ExcelFileGenMixin, TestCase):

    def setUp(self):
        XLSXDeserializer = serializers.get_deserializer("xlsx")
        self.xlsx_deserializer = XLSXDeserializer(self._get_xlsx_stream())

    def test_objects_must_be_saved(self):
        list(self.xlsx_deserializer)

        # All the objects in the first sheet must be saved:
        self.assertEqual(3, Person.objects.all().count())
        self.assertEquals(1, Person.objects.first().id)
        self.assertEquals('Person 1', Person.objects.first().name)
        self.assertEquals(21, Person.objects.first().age)

        # All the objects in the second sheet must be saved:
        self.assertEqual(4, Ingredient.objects.all().count())
        self.assertEquals(1, Ingredient.objects.first().id)
        self.assertEquals('Ingredient 1', Ingredient.objects.first().name)

        # All the objects in the second sheet must be saved:
        self.assertEqual(2, Recipe.objects.all().count())
        self.assertEquals(1, Recipe.objects.first().id)
        self.assertEquals('Recipe 1', Recipe.objects.first().name)
        self.assertEquals(1, Recipe.objects.first().owner.id)


class XlsxDeserializerUnitTest(ExcelFileGenMixin, TestCase):

    def setUp(self):
        XLSXDeserializer = serializers.get_deserializer("xlsx")
        self.xlsx_deserializer = XLSXDeserializer(self._get_xlsx_stream())

    def test_get_model(self):
        workbook = self.xlsx_deserializer.workbook
        sheet1 = workbook['myapp.Person']
        sheet2 = workbook['myapp.Ingredient']

        sheet1_model_class = self.xlsx_deserializer._get_model(sheet1)
        self.assertEquals(sheet1_model_class, Person)

        sheet2_model_class = self.xlsx_deserializer._get_model(sheet2)
        self.assertEquals(sheet2_model_class, Ingredient)

    def test_get_model_invalid_model(self):
        """ get_model should raise an exception if model identifier is invalid. """
        workbook = self.xlsx_deserializer.workbook
        sheet1 = workbook['myapp.Person']
        sheet1.title = 'Invalid.Model.Name'
        with self.assertRaises(Exception) as context:
            self.xlsx_deserializer._get_model(sheet1)

    def test_row_to_dict(self):
        """
        row_to_dict receives a sheet row and fields list (base on the sheet
        header AND the obj model) and returns a high level python obj.
        """
        # select_sheet 1:
        sheet1_fields = [
            ('id', Person._meta.get_field('id')),
            ('name', Person._meta.get_field('name')),
            ('age', Person._meta.get_field('age')),
        ]
        sheet1_first_obj_row = self.xlsx_deserializer.workbook['myapp.Person'][2]
        result = self.xlsx_deserializer._row_to_dict(
            sheet1_first_obj_row,
            sheet1_fields
        )
        expected_result = {
            'id': 1,
            'name': 'Person 1',
            'age': '21'
        }
        self.assertEquals(expected_result, result)

        # select_sheet 2:
        sheet2_fields = [
            ('id', Ingredient._meta.get_field('id')),
            ('name', Ingredient._meta.get_field('name')),
        ]
        sheet2_first_obj_row = self.xlsx_deserializer.workbook['myapp.Ingredient'][2]
        result = self.xlsx_deserializer._row_to_dict(
            sheet2_first_obj_row,
            sheet2_fields
        )
        expected_result = {
            'id': 1,
            'name': 'Ingredient 1'
        }
        self.assertEquals(expected_result, result)

    def test_current_row_is_valid(self):
        """ current_row is valid if it is pointing to a row with data. """

        # select_sheet 1 - current_row valid between 2 - 4
        self.xlsx_deserializer._select_sheet('myapp.Person')
        self.assertTrue(self.xlsx_deserializer._current_row_is_valid())
        self.xlsx_deserializer.current_row = 2
        self.assertTrue(self.xlsx_deserializer._current_row_is_valid())
        self.xlsx_deserializer.current_row = 3
        self.assertTrue(self.xlsx_deserializer._current_row_is_valid())
        self.xlsx_deserializer.current_row = 4
        self.assertTrue(self.xlsx_deserializer._current_row_is_valid())
        self.xlsx_deserializer.current_row = 5
        self.assertFalse(self.xlsx_deserializer._current_row_is_valid())

        # select_sheet 2 - current_row valid between 2 - 5
        self.xlsx_deserializer._select_sheet('myapp.Ingredient')
        self.assertTrue(self.xlsx_deserializer._current_row_is_valid())
        self.xlsx_deserializer.current_row = 2
        self.assertTrue(self.xlsx_deserializer._current_row_is_valid())
        self.xlsx_deserializer.current_row = 5
        self.assertTrue(self.xlsx_deserializer._current_row_is_valid())
        self.xlsx_deserializer.current_row = 6
        self.assertFalse(self.xlsx_deserializer._current_row_is_valid())

        # select_sheet 3
        self.xlsx_deserializer._select_sheet('myapp.Recipe')
        self.assertTrue(self.xlsx_deserializer._current_row_is_valid())
        self.xlsx_deserializer.current_row = 6
        self.assertFalse(self.xlsx_deserializer._current_row_is_valid())

    def test_reset_current_row(self):
        """
        select_sheet should reset current row to the first content line,
        which is the second line (2) bc the first line is the sheet's header.
        """
        self.assertEquals(2, self.xlsx_deserializer.current_row)
        self.xlsx_deserializer._reset_current_row()
        self.assertEquals(2, self.xlsx_deserializer.current_row)

        self.xlsx_deserializer.current_row = 5
        self.xlsx_deserializer._reset_current_row()
        self.assertEquals(2, self.xlsx_deserializer.current_row)

        self.xlsx_deserializer.current_row = 10
        self.xlsx_deserializer._reset_current_row()
        self.assertEquals(2, self.xlsx_deserializer.current_row)

    def test_has_next_sheet(self):
        """ _has_next_sheet returns if there is a next sheet to be activate. """
        workbook = self.xlsx_deserializer.workbook
        sheet1 = workbook['myapp.Person']
        sheet2 = workbook['myapp.Ingredient']
        sheet3 = workbook['myapp.Recipe']
        sheet4 = workbook['myapp.Association']

        # Activate sheet 1:
        workbook.active = sheet1
        self.assertTrue(self.xlsx_deserializer._has_next_sheet())

        #Activate sheet 2:
        workbook.active = sheet2
        self.assertTrue(self.xlsx_deserializer._has_next_sheet())

        #Activate sheet 3:
        workbook.active = sheet3
        self.assertTrue(self.xlsx_deserializer._has_next_sheet())

        #Activate sheet 4:
        workbook.active = sheet4
        self.assertFalse(self.xlsx_deserializer._has_next_sheet())

    def test_select_first_sheet(self):
        """ select_first_sheet should select the first sheet. """
        workbook = self.xlsx_deserializer.workbook
        sheet1 = workbook['myapp.Person']
        sheet2 = workbook['myapp.Ingredient']

        # Activate sheet 2:
        workbook.active = sheet2
        self.assertEqual(sheet2, workbook.active)
        self.assertEqual('myapp.Ingredient', workbook.active.title)

        # select_first_sheet:
        self.xlsx_deserializer._select_first_sheet()
        self.assertEqual(sheet1, workbook.active)
        self.assertEqual('myapp.Person', workbook.active.title)

    def test_select_next_sheet(self):
        """ select_next_sheet should select the next sheet. """
        workbook = self.xlsx_deserializer.workbook
        sheet1 = workbook['myapp.Person']
        sheet2 = workbook['myapp.Ingredient']

        # Activate sheet 1:
        workbook.active = sheet1
        self.assertEqual(sheet1, workbook.active)
        self.assertEqual('myapp.Person', workbook.active.title)

        # select_next_sheet:
        self.xlsx_deserializer._select_next_sheet()
        self.assertEqual(sheet2, workbook.active)
        self.assertEqual('myapp.Ingredient', workbook.active.title)

    def test_select_next_sheet_exception(self):
        """ change_to_next_sheet must raise exception when last sheet is selected"""
        workbook = self.xlsx_deserializer.workbook
        sheet3 = workbook['myapp.Association']

        # Activate sheet 2 (no next sheet):
        workbook.active = sheet3
        with self.assertRaises(Exception) as context:
            self.xlsx_deserializer._select_next_sheet()

    def test_select_next_sheet_should_call_start_sheet(self):
        """ change_to_next_sheet should start the next sheet."""
        workbook = self.xlsx_deserializer.workbook
        sheet1 = workbook['myapp.Person']
        sheet2 = workbook['myapp.Ingredient']

        # Activate sheet 1:
        workbook.active = sheet1
        with patch('django_excel_fixture.serializers.xlsx_serializer.Deserializer._start_sheet') as start_sheet:
            self.xlsx_deserializer._select_next_sheet()
            start_sheet.assert_called()

    def test_select_sheet_update_current_model(self):
        """
        select_sheet should update current model.
        Every sheet is associated with one model,
        when select_sheet is called, the current model
        should be updated for the new model.
        """

        # select_sheet 1:
        self.xlsx_deserializer._select_sheet('myapp.Person')
        self.assertEquals(Person, self.xlsx_deserializer.model_class)

        # select_sheet 2:
        self.xlsx_deserializer._select_sheet('myapp.Ingredient')
        self.assertEquals(Ingredient, self.xlsx_deserializer.model_class)

    def test_select_sheet_update_active_sheet(self):
        """ select_sheet should activate new sheet. """

        # select_sheet 1:
        self.xlsx_deserializer._select_sheet('myapp.Person')
        self.assertEquals('myapp.Person', self.xlsx_deserializer.workbook.active.title)

        # select_sheet 2:
        self.xlsx_deserializer._select_sheet('myapp.Ingredient')
        self.assertEquals('myapp.Ingredient', self.xlsx_deserializer.workbook.active.title)

    def test_select_sheet_update_num_filed_and_num_objects(self):
        """ select_sheet should update number of fields and number of objects."""

        # select_sheet 1:
        self.xlsx_deserializer._select_sheet('myapp.Person')
        self.assertEquals(3, self.xlsx_deserializer.num_fields)
        self.assertEquals(3, self.xlsx_deserializer.num_objects)

        # select_sheet 2:
        self.xlsx_deserializer._select_sheet('myapp.Ingredient')
        self.assertEquals(2, self.xlsx_deserializer.num_fields)
        self.assertEquals(4, self.xlsx_deserializer.num_objects)

    def test_select_sheet_reset_current_row(self):
        """ select_sheet should reset current row to second line (2)."""

        # select_sheet 1:
        self.xlsx_deserializer.current_row = 10
        self.xlsx_deserializer._select_sheet('myapp.Person')
        self.assertEquals(2, self.xlsx_deserializer.current_row)

        # select_sheet 2:
        self.xlsx_deserializer.current_row = 10
        self.xlsx_deserializer._select_sheet('myapp.Ingredient')
        self.assertEquals(2, self.xlsx_deserializer.current_row)

    # Missing fields test:
    def test_select_sheet_update_model_fields(self):
        """
        select_sheet should update fields list and model_fields dict.
        Every sheet is associated with one model,
        when select_sheet is called, the fields variable
        should have a dict of fields of the current sheet
        (based on the header and the model).

        model_fields:
            -Is a dict.
            -Based on the model's fields.
            -All the elements on the model is an item.

        fields:
            -Is a list of tuple.
            -Based on the sheet's header.
            -All the elements on the current sheet's header is an item.
            -Can have less item than the model_fields.

        Example:
            model = myapp.Person

            model_fields = {
                'id': < django.db.models.fields.AutoField: id >,
                'name': < django.db.models.fields.CharField: name >,
                'age': < django.db.models.fields.IntegerField: age >,
            }

            fields = [
                ('id', <django.db.models.fields.AutoField: id>),
                ('name', <django.db.models.fields.CharField: name>),
                ('age', <django.db.models.fields.IntegerField: age>),
            ]

            * Fields can have less items than model_fields
              (depending on the sheet's header).

        """
        # select_sheet 1:
        self.xlsx_deserializer._select_sheet('myapp.Person')
        for key in ['id', 'name', 'age']:
            self.assertIn(key, self.xlsx_deserializer.model_fields)

        self.assertIsInstance(
            self.xlsx_deserializer.model_fields['id'],
            models.fields.AutoField
        )
        self.assertIsInstance(
            self.xlsx_deserializer.model_fields['name'],
            models.fields.CharField
        )
        self.assertIsInstance(
            self.xlsx_deserializer.model_fields['age'],
            models.fields.IntegerField
        )

        """
        {
            'id': < django.db.models.fields.AutoField: id >, 
            'name': < django.db.models.fields.CharField: name >, 
            'age': < django.db.models.fields.IntegerField: age >,
        }
        """

        # select_sheet 2:
        self.xlsx_deserializer._select_sheet('myapp.Ingredient')
        for key in ['id', 'name']:
            self.assertIn(key, self.xlsx_deserializer.model_fields)

        # Just to be sure, since both models are very close to each other:
        self.assertNotIn('age', self.xlsx_deserializer.model_fields)

        self.assertIsInstance(
            self.xlsx_deserializer.model_fields['id'],
            models.fields.AutoField
        )
        self.assertIsInstance(
            self.xlsx_deserializer.model_fields['name'],
            models.fields.CharField
        )

        """
        {
            'id': <django.db.models.fields.AutoField: id>, 
            'name': <django.db.models.fields.CharField: name>
        }
        """

    def test_select_sheet_update_fields(self):
        """ select_sheet should update fields dict."""

        # select_sheet 1:
        self.xlsx_deserializer._select_sheet('myapp.Person')
        sheet1_fields = [
            ('id', Person._meta.get_field('id')),
            ('name', Person._meta.get_field('name')),
            ('age', Person._meta.get_field('age')),
        ]
        self.assertEquals(
            sheet1_fields,
            self.xlsx_deserializer.fields
        )

        # select_sheet 2:
        self.xlsx_deserializer._select_sheet('myapp.Ingredient')
        sheet2_fields = [
            ('id', Ingredient._meta.get_field('id')),
            ('name', Ingredient._meta.get_field('name')),
        ]
        self.assertEquals(
            sheet2_fields,
            self.xlsx_deserializer.fields
        )

    def test_get_value_ForeignKey(self):

        # Object must exist:
        Person.objects.create(name='Person 1', age=21)
        Person.objects.create(name='Person 2', age=22)

        # Create field:
        class TestModelForeignKey(models.Model):
            test_field = models.ForeignKey('Person', blank=True, null=True, on_delete=models.SET_NULL)
        field = TestModelForeignKey._meta.get_field('test_field')

        cell = Workbook().active.cell(row=1, column=1, value=1)
        value = self.xlsx_deserializer._get_value(cell, field)
        self.assertIsInstance(value, Person)
        self.assertEquals(value.id, 1)

        cell = Workbook().active.cell(row=1, column=1, value='2')
        value = self.xlsx_deserializer._get_value(cell, field)
        self.assertIsInstance(value, Person)
        self.assertEquals(value.id, 2)

        cell = Workbook().active.cell(row=1, column=1, value=None)
        value = self.xlsx_deserializer._get_value(cell, field)
        self.assertIsNone(value)

    def test_get_value_DateTime(self):

        class TestModelDateTime(models.Model):
            test_field = models.DateTimeField(auto_now_add=True)
        field = TestModelDateTime._meta.get_field('test_field')

        cell = Workbook().active.cell(row=1, column=1, value=datetime.datetime.now())
        value = self.xlsx_deserializer._get_value(cell, field)
        self.assertIsInstance(value, datetime.datetime)

        cell = Workbook().active.cell(row=2, column=1, value='2019-09-30 13:03:56:307439')
        value = self.xlsx_deserializer._get_value(cell, field)
        self.assertIsInstance(value, datetime.datetime)

    def test_get_value_CharField(self):

        class TestModelChar(models.Model):
            test_field = models.CharField(blank=True, null=True, max_length=5)
        field = TestModelChar._meta.get_field('test_field')

        cell = Workbook().active.cell(row=1, column=1, value='abcde')
        value = self.xlsx_deserializer._get_value(cell, field)
        self.assertIsInstance(value, str)
        self.assertEqual(value, 'abcde')

        # get_value does not validate. This will raise when being saved:
        cell = Workbook().active.cell(row=1, column=1, value='abcdef')
        value = self.xlsx_deserializer._get_value(cell, field)
        self.assertIsInstance(value, str)
        self.assertEqual(value, 'abcdef')
