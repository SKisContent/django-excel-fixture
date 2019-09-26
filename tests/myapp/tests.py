# coding: utf-8
from django.db import models
from io import BytesIO
from unittest import skip
from unittest.mock import patch, MagicMock
from django.core import serializers
from django.test import TestCase
from openpyxl import Workbook
from openpyxl.writer.excel import save_workbook
from .models import Person, Recipe, Ingredient, Association


class XlsxSerializerTest(TestCase):
    fixtures = ['fixture.json']

    @classmethod
    def setUpTestData(cls):
        pass

    def setUp(self):
        pass

    def test_serializer_has_been_registered(self):
        public_formats = serializers.get_public_serializer_formats()
        self.assertIn('xlsx', public_formats)

    def _get_field_values(self, stream, field_name):
        pass

    def _get_pk_value(self, stream):
        pass

    def _validate_output(self, stream):
        pass


class XlsxSerializerUnitTest(TestCase):
    fixtures = ['fixture.json']

    def setUp(self):
        XLSXSerializer = serializers.get_serializer("xlsx")
        self.xlsx_serializer = XLSXSerializer()
        # Mocking:
        self.xlsx_serializer.use_natural_primary_keys = False

    def test_start_serialization(self):
        """ A new serializer should have a workbook. """
        self.assertIsNone(self.xlsx_serializer.workbook)
        self.xlsx_serializer.start_serialization()
        self.assertIsInstance(self.xlsx_serializer.workbook, Workbook)

    def test_after_start_number_of_sheet(self):
        """ A new serializer should have no sheet. """
        self.xlsx_serializer.start_serialization()
        self.assertEquals(0, len(self.xlsx_serializer.workbook.worksheets))

    def _start_object(self):
        self.obj = Person(name='Henrique Portela', age=41)
        self.xlsx_serializer.start_serialization()
        self.xlsx_serializer.start_object(self.obj)

    def test_start_object_create_sheet(self):
        """ If the obj sheet doesn't exist, it must be created. """
        self._start_object()
        self.assertEquals(1, len(self.xlsx_serializer.workbook.worksheets))

    def test_start_object_sheet_name(self):
        """ Name of the sheet, should be obj class. """
        self._start_object()
        self.assertIn('myapp.Person', self.xlsx_serializer.workbook.sheetnames)

    def test_start_object_sheet_header(self):
        """ Newly created sheet, must have header. """
        self._start_object()
        self.assertEquals('id' , self.xlsx_serializer.workbook['myapp.Person']['A1'].value)
        self.assertEquals('name' , self.xlsx_serializer.workbook['myapp.Person']['B1'].value)
        self.assertEquals('age' , self.xlsx_serializer.workbook['myapp.Person']['C1'].value)

    def test_sheet_header_format(self):
        """ The header should be formated. """
        self._start_object()
        self.assertEquals('Calibri', self.xlsx_serializer.workbook['myapp.Person']['A1'].font.name)
        self.assertEquals(11, self.xlsx_serializer.workbook['myapp.Person']['A1'].font.size)
        self.assertEquals(True, self.xlsx_serializer.workbook['myapp.Person']['A1'].font.bold)
        self.assertEquals('solid', self.xlsx_serializer.workbook['myapp.Person']['A1'].fill.patternType)
        self.assertEquals('009BBB59', self.xlsx_serializer.workbook['myapp.Person']['A1'].fill.fgColor.rgb)
        self.assertEquals('center', self.xlsx_serializer.workbook['myapp.Person']['A1'].alignment.horizontal)
        self.assertEquals('center', self.xlsx_serializer.workbook['myapp.Person']['A1'].alignment.vertical)

    def test_start_object_current_row(self):
        """ After creating a new sheet, current row should be pointing to line 2. """
        self._start_object()
        self.assertEquals(2, self.xlsx_serializer.current_row)

    def test_start_object_curret_row_update(self):
        """ After start object, current_row should be incremented. """
        self._start_object()
        for i in range(3,10):
            self.obj = Person(name='Henrique Portela', age=41)
            self.xlsx_serializer.start_object(self.obj)
            self.assertEquals(i, self.xlsx_serializer.current_row)

    def test_start_object_active_sheet(self):
        """ When second obj is started, active sheet should change"""
        self._start_object()
        self.obj = Association.objects.first()
        self.xlsx_serializer.start_object(self.obj)
        last_sheet = self.xlsx_serializer.workbook['myapp.Association']
        self.assertEquals(last_sheet, self.xlsx_serializer.workbook.active)

    def test_start_object_add_pk(self):
        """ start_object should add primary key when NOT natural primary_keys. """
        self._start_object()
        for i in range(3, 10):
            obj = Person(name='Henrique Portela', age=41)
            obj.save()
            self.xlsx_serializer.start_object(obj)
            self.assertEquals(obj.pk, self.xlsx_serializer.workbook['myapp.Person'].cell(row=i, column=1).value)

    def test_handle_field(self):
        """ handle_field should add the value to the correct cell. """
        self._start_object()

        self.xlsx_serializer.handle_field(self.obj, self.obj._meta.fields[1])
        self.assertEquals('Henrique Portela', self.xlsx_serializer.workbook['myapp.Person']['B2'].value)
        self.xlsx_serializer.handle_field(self.obj, self.obj._meta.fields[2])
        self.assertEquals(41, self.xlsx_serializer.workbook['myapp.Person']['C2'].value)

        self.obj = Person(name='John Smith', age=51)
        self.xlsx_serializer.start_object(self.obj)

        self.xlsx_serializer.handle_field(self.obj, self.obj._meta.fields[1])
        self.assertEquals('John Smith', self.xlsx_serializer.workbook['myapp.Person']['B3'].value)
        self.xlsx_serializer.handle_field(self.obj, self.obj._meta.fields[2])
        self.assertEquals(51, self.xlsx_serializer.workbook['myapp.Person']['C3'].value)

    @skip('DateTimeField without value, should now crash')
    def test_handle_field_empty(self):
        pass

    def test_column_index(self):
        """ _column_index should return the index from field value. """
        self._start_object()
        self.assertEquals(1, self.xlsx_serializer._column_index('id'))
        self.assertEquals(2, self.xlsx_serializer._column_index('name'))
        self.assertEquals(3, self.xlsx_serializer._column_index('age'))

    def test_handle_fk_field(self):
        """ handle_fk_field should add the value to the correct cell. """
        self.obj = Association.objects.first()
        self.xlsx_serializer.start_serialization()
        self.xlsx_serializer.start_object(self.obj)

        self.xlsx_serializer.handle_fk_field(self.obj, self.obj._meta.fields[1])
        self.assertEquals(1, self.xlsx_serializer.workbook['myapp.Association']['B2'].value)
        self.xlsx_serializer.handle_fk_field(self.obj, self.obj._meta.fields[2])
        self.assertEquals(1, self.xlsx_serializer.workbook['myapp.Association']['C2'].value)


class XlsxDeserializerUnitTest(TestCase):

    def _get_xlsx_stream(self):
        workbook = Workbook()

        # Remove auto-created sheet:
        name = workbook.get_sheet_names()[0]
        sheet = workbook.get_sheet_by_name(name)
        workbook.remove_sheet(sheet)

        # Add first model sheet:
        sheet1 = workbook.create_sheet('myapp.Person')
        workbook.active = sheet1 # Don't know if this line is needed.
        # Add header:
        workbook['myapp.Person'].cell(row=1, column=1, value='id')
        workbook['myapp.Person'].cell(row=1, column=2, value='name')
        workbook['myapp.Person'].cell(row=1, column=3, value='age')
        # Add content 1:
        workbook['myapp.Person'].cell(row=2, column=1, value=1)
        workbook['myapp.Person'].cell(row=2, column=2, value='Person 1')
        workbook['myapp.Person'].cell(row=2, column=3, value=21)
        # Add content 2:
        workbook['myapp.Person'].cell(row=3, column=1, value=2)
        workbook['myapp.Person'].cell(row=3, column=2, value='Person 2')
        workbook['myapp.Person'].cell(row=3, column=3, value=22)
        # Add content 3:
        workbook['myapp.Person'].cell(row=4, column=1, value=3)
        workbook['myapp.Person'].cell(row=4, column=2, value='Person 3')
        workbook['myapp.Person'].cell(row=4, column=3, value=23)

        # Add second model sheet:
        sheet1 = workbook.create_sheet('myapp.Ingredient')
        workbook.active = sheet1 # Don't know if this line is needed.
        # Add header:
        workbook['myapp.Ingredient'].cell(row=1, column=1, value='id')
        workbook['myapp.Ingredient'].cell(row=1, column=2, value='name')
        # Add content 1:
        workbook['myapp.Ingredient'].cell(row=2, column=1, value=1)
        workbook['myapp.Ingredient'].cell(row=2, column=2, value='Ingredient 1')
        # Add content 2:
        workbook['myapp.Ingredient'].cell(row=3, column=1, value=2)
        workbook['myapp.Ingredient'].cell(row=3, column=2, value='Ingredient 2')
        # Add content 3:
        workbook['myapp.Ingredient'].cell(row=4, column=1, value=3)
        workbook['myapp.Ingredient'].cell(row=4, column=2, value='Ingredient 3')
        # Add content 4:
        workbook['myapp.Ingredient'].cell(row=5, column=1, value=4)
        workbook['myapp.Ingredient'].cell(row=5, column=2, value='Ingredient 4')

        # Generate in memory stream:
        # About StringIO:
        #       Implements a file-like class that reads and writes a string
        #       buffer (also known as memory files).
        in_memory_file = BytesIO()
        save_workbook(workbook, in_memory_file)

        # Saving the file to inspect:
        #save_workbook(workbook, 'z_gen_file_01.xlsx')

        return in_memory_file

    @classmethod
    def setUpTestData(cls):
        pass

    def setUp(self):
        XLSXDeserializer = serializers.get_deserializer("xlsx")
        self.xlsx_deserializer = XLSXDeserializer(self._get_xlsx_stream())

    @skip('Just for now.')
    def test_objects_must_be_saved(self):
        list(self.xlsx_deserializer)

        # All the objects in the first sheet must be saved:
        self.assertEqual(3, Person.objects.all().count())
        self.assertEquals(1, Person.objects.first().id)
        self.assertEquals('Person 1', Person.objects.first().name)
        self.assertEquals(21, Person.objects.first().age)

        # All the objects in the second sheet must be saved:
        self.assertEqual(4, Ingredient.objects.all().count())
        self.assertEquals(1, Ingredient.objects.first().id)
        self.assertEquals('Ingredient 1', Ingredient.objects.first().name)

    def test_has_next_sheet(self):
        workbook = self.xlsx_deserializer.workbook
        sheet1 = workbook['myapp.Person']
        sheet2 = workbook['myapp.Ingredient']

        # Activate sheet 1:
        workbook.active = sheet1
        self.assertTrue(self.xlsx_deserializer._has_next_sheet())

        #Activate sheet 2:
        workbook.active = sheet2
        self.assertFalse(self.xlsx_deserializer._has_next_sheet())

    def test_select_next_sheet(self):
        workbook = self.xlsx_deserializer.workbook
        sheet1 = workbook['myapp.Person']
        sheet2 = workbook['myapp.Ingredient']

        # Activate sheet 1:
        workbook.active = sheet1
        self.assertEqual(sheet1, workbook.active)
        self.assertEqual('myapp.Person', workbook.active.title)

        # Change_to_next_sheet:
        self.xlsx_deserializer._select_next_sheet()
        self.assertEqual(sheet2, workbook.active)
        self.assertEqual('myapp.Ingredient', workbook.active.title)

    def test_select_next_sheet_exception(self):
        """ change_to_next_sheet must raise exception when in last sheet"""
        workbook = self.xlsx_deserializer.workbook
        sheet1 = workbook['myapp.Person']
        sheet2 = workbook['myapp.Ingredient']

        # Activate sheet 2 (no next sheet):
        workbook.active = sheet2
        with self.assertRaises(Exception) as context:
            self.xlsx_deserializer._select_next_sheet()

    def test_select_next_sheet_should_call_start_sheet(self):
        """ change_to_next_sheet should start the next sheet."""
        workbook = self.xlsx_deserializer.workbook
        sheet1 = workbook['myapp.Person']
        sheet2 = workbook['myapp.Ingredient']

        # Activate sheet 1:
        workbook.active = sheet1

        # Mocking:
        with patch('django_excel_fixture.serializers.xlsx_serializer.Deserializer._start_sheet') as start_sheet:
            self.xlsx_deserializer._select_next_sheet()
            start_sheet.assert_called()
            # Refactoring start_sheet to work with active sheet
            # start_sheet.assert_called_with('myapp.Ingredient')

    #
    # Refactoring tests to call start_sheet indirectly
    #

    def test_start_sheet_update_current_model(self):
        """ start_sheet should update current model."""
        workbook = self.xlsx_deserializer.workbook
        sheet1 = workbook['myapp.Person']
        sheet2 = workbook['myapp.Ingredient']

        # start_sheet 1:
        self.xlsx_deserializer._select_sheet('myapp.Person')
        self.assertEquals(Person, self.xlsx_deserializer.model_class)

        # start_sheet 2:
        self.xlsx_deserializer._select_sheet('myapp.Ingredient')
        self.assertEquals(Ingredient, self.xlsx_deserializer.model_class)

    @skip('not model_field but fields')
    def test_start_sheet_update_model_fields_dict(self):
        """ start_sheet should update field list. """
        workbook = self.xlsx_deserializer.workbook
        sheet1 = workbook['myapp.Person']
        sheet2 = workbook['myapp.Ingredient']

        # start_sheet 1:
        self.xlsx_deserializer._select_sheet('myapp.Person')
        for key in ['id', 'name', 'age']:
            self.assertIn(key, self.xlsx_deserializer.model_fields)

        self.assertIsInstance(
            self.xlsx_deserializer.model_fields['id'],
            models.fields.AutoField
        )
        self.assertIsInstance(
            self.xlsx_deserializer.model_fields['name'],
            models.fields.CharField
        )
        self.assertIsInstance(
            self.xlsx_deserializer.model_fields['age'],
            models.fields.IntegerField
        )
        """
        {
            'id': < django.db.models.fields.AutoField: id >, 
            'name': < django.db.models.fields.CharField: name >, 
            'age': < django.db.models.fields.IntegerField: age >,
        }
        """

        # start_sheet 2:
        self.xlsx_deserializer._select_sheet('myapp.Ingredient')
        for key in ['id', 'name']:
            self.assertIn(key, self.xlsx_deserializer.model_fields)

        # Just to be sure, since both models are very close to each other:
        self.assertNotIn('age', self.xlsx_deserializer.model_fields)

        self.assertIsInstance(
            self.xlsx_deserializer.model_fields['id'],
            models.fields.AutoField
        )
        self.assertIsInstance(
            self.xlsx_deserializer.model_fields['name'],
            models.fields.CharField
        )

        """
        {
            'id': <django.db.models.fields.AutoField: id>, 
            'name': <django.db.models.fields.CharField: name>
        }
        """

    def test_start_sheet_update_active_sheet(self):
        """ start_sheet should activate new sheet."""
        workbook = self.xlsx_deserializer.workbook
        sheet1 = workbook['myapp.Person']
        sheet2 = workbook['myapp.Ingredient']

        # start_sheet 1:
        self.xlsx_deserializer._select_sheet('myapp.Person')
        self.assertEquals('myapp.Person', self.xlsx_deserializer.workbook.active.title)

        # start_sheet 2:
        self.xlsx_deserializer._select_sheet('myapp.Ingredient')
        self.assertEquals('myapp.Ingredient', self.xlsx_deserializer.workbook.active.title)

    def test_start_sheet_update_num_filed_and_num_objects(self):
        """ start_sheet should update number of fields and number of objects."""
        workbook = self.xlsx_deserializer.workbook
        sheet1 = workbook['myapp.Person']
        sheet2 = workbook['myapp.Ingredient']

        # start_sheet 1:
        self.xlsx_deserializer._select_sheet('myapp.Person')
        self.assertEquals(3, self.xlsx_deserializer.num_fields)
        self.assertEquals(3, self.xlsx_deserializer.num_objects)

        # start_sheet 2:
        self.xlsx_deserializer._select_sheet('myapp.Ingredient')
        self.assertEquals(2, self.xlsx_deserializer.num_fields)
        self.assertEquals(4, self.xlsx_deserializer.num_objects)

    def test_start_sheet_should_reset_current_row(self):
        """ start_sheet should reset current row to second line (2)."""
        workbook = self.xlsx_deserializer.workbook
        sheet1 = workbook['myapp.Person']
        sheet2 = workbook['myapp.Ingredient']

        # start_sheet 1:
        self.xlsx_deserializer.current_row = 10
        self.xlsx_deserializer._select_sheet('myapp.Person')
        self.assertEquals(2, self.xlsx_deserializer.current_row)

        # start_sheet 2:
        self.xlsx_deserializer.current_row = 10
        self.xlsx_deserializer._select_sheet('myapp.Ingredient')
        self.assertEquals(2, self.xlsx_deserializer.current_row)
