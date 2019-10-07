"""
Serializes a Django Model to Excel and vice-versa
Some observations:
OpenPyXL offers both reference addressing (e.g., 'A1', 'AC93') and indexed addressing.
When indexing into a row, the base (first) index is 1. However, after that you're
in the world of python tuples, so the indexes are 0-based (e.g., ws[1][0] == ws['A1'], ws[2][4] == ws['E2'])
The same way, columns can be directly referenced (ws['A']), but then rows are 0-based
(e.g., ws['A'][0] == ws['A1'], ws['C'][4] == ws['C5'])
"""
import decimal
import django
import io
import re
from datetime import datetime, date

import pytz
from django.apps import apps
from django.conf import settings
from django.core.serializers import base
from django.utils import timezone
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.writer.excel import save_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.styles import colors, borders, fills

from django.db import models
from django.db.models import AutoField, BooleanField, CharField, DurationField, DateTimeField, ForeignKey, DecimalField, \
    IntegerField, PositiveIntegerField, BigAutoField, BigIntegerField, DateField, EmailField, FileField, FilePathField, \
    FloatField, ImageField, GenericIPAddressField, NullBooleanField, PositiveSmallIntegerField, SlugField, \
    SmallIntegerField, TextField, TimeField, URLField, UUIDField


PREFERRED_TS_FORMAT = '%Y-%m-%d %H:%M:%S:%f%z'
Y_M_D_FORMAT = '%Y-%m-%d'

DATETIME_FORMATS = {r'\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}.\d{6}[-+]\d{4}':PREFERRED_TS_FORMAT,
                    r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}.\d{6}':'%Y-%m-%d %H:%M:%S:%f',
                    r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}':'%Y-%m-%d %H:%M:%S',
                    }

HEADER_FORMAT = {
    'font': Font(
        name='Calibri',
        size=11,
        bold=True,
    ),
    'fill': PatternFill(
        fill_type='solid',
        fgColor='9BBB59',
    ),
    'alignment': Alignment(
        horizontal='center',
        vertical='center',
    ),
}


class Serializer(base.Serializer):
    """
        Serializer a QuerySet to a xlsx file format.
            - Every model class will be a sheet.
            - Every field will be a column.
            - Every obj will be a row.
    """

    workbook = None

    COMPATIBLE_FIELDS = [
        AutoField,
        BigAutoField,
        BigIntegerField,
        BooleanField,
        CharField,
        DateField,
        DateTimeField,
        DecimalField,
        DurationField,
        EmailField,
        FileField,
        FilePathField,
        FloatField,
        ImageField,
        IntegerField,
        GenericIPAddressField,
        NullBooleanField,
        PositiveIntegerField,
        PositiveSmallIntegerField,
        SlugField,
        SmallIntegerField,
        TextField,
        TimeField,
        URLField,
        UUIDField,

        ForeignKey,
        # ManyToManyField,
        # OneToOneField,
    ]

    def start_serialization(self):
        """
        Start Serialization -- Create workbook
        """
        self.workbook = Workbook()
        name = self.workbook.get_sheet_names()[0]
        sheet = self.workbook.get_sheet_by_name(name)
        self.workbook.remove_sheet(sheet)

    def start_object(self, obj):
        """
        Start Object -- Create new sheet, if necessary
        """

        if not hasattr(obj, "_meta"):
            raise base.SerializationError("Non-model object (%s) encountered during serialization" % type(obj))

        sheet_name = obj._meta.label
        if sheet_name not in self.workbook.get_sheet_names():
            sheet = self.workbook.create_sheet(sheet_name) # Create sheet in the last position
            self.workbook.active = sheet

            # Create header:
            for index, field in enumerate(obj._meta.fields):
                self.workbook[sheet_name].cell(row=1, column=(index+1), value=field.name)
                self.workbook[sheet_name].cell(row=1, column=(index+1)).font = HEADER_FORMAT['font']
                self.workbook[sheet_name].cell(row=1, column=(index+1)).fill = HEADER_FORMAT['fill']
                self.workbook[sheet_name].cell(row=1, column=(index+1)).alignment = HEADER_FORMAT['alignment']

            # Set current row
            self.current_row = 1

        # Increase row cursor:
        self.current_row += 1

        # Add primary key:
        if not self.use_natural_primary_keys or not hasattr(obj, 'natural_key'):
            obj_pk = obj.pk
            if obj_pk is not None:
                self.workbook[sheet_name].cell(row=self.current_row, column=1, value=obj_pk)

    def _column_index(self, field_name):
        """ Return the column index for the ACTIVE sheet"""

        for i in range(1, self.workbook.active.max_column+1):
            if field_name == self.workbook.active.cell(row=1, column=i).value:
                return i

        raise Exception('Field name "{}" not found in sheet'.format(field_name))

    def handle_field(self, obj, field):

        try:

            if isinstance(field, models.DateTimeField):
                value = field.value_from_object(obj)
                if value:
                    value = value.strftime(PREFERRED_TS_FORMAT)

            elif isinstance(field, models.ImageField):
                value = field.value_to_string(obj)

            elif type(field) in self.COMPATIBLE_FIELDS:
                value = field.value_from_object(obj)

            else:
                value = field.value_to_string(obj)

            self.workbook.active.cell(
                row=self.current_row,
                column=self._column_index(field.name),
                value=value
            )

        except base.SerializationError:
            raise ValueError("%s.%s (pk:%s) contains unserializable characters" %
                             (obj.__class__.__name__, field.name, obj.pk))

    def handle_fk_field(self, obj, field):

        value = field.value_from_object(obj)

        self.workbook.active.cell(
            row=self.current_row,
            column=self._column_index(field.name),
            value=value
        )



    def handle_m2m_field(self, obj, field):
        pass

    def end_serialization(self):

        filename = self.stream.name
        if filename == '<stdout>':
            # If there is no file, dump a CSV representation to stdout
            out_str = self.csv(self.workbook.active)
            self.stream.write(out_str)
        else:
            # The default stream is opened in text mode, but we need binary
            self.stream.close()
            # Just use the openpyxl saving method
            save_workbook(self.workbook, filename)


    def csv(self, worksheet):
        num_rows = len(worksheet['A'])
        if num_rows == 1:
            raise base.SerializationError('There is no data to dump.')
        buffer = io.StringIO()
        for ix in range(num_rows):
            buffer.write(','.join([wrap(cell.value) if cell.value else '' for cell in worksheet[ix+1]]))
            buffer.write('\n')
        return buffer.getvalue()


def wrap(value):
    if value is None:
        return None
    if type(value) is str:
        value.replace('"','""')
    else:
        value = str(value)
    return '"{0}"'.format(value)




class Deserializer(base.Deserializer):
    """
    openpyxl:
    workbook.sheetnames  or  workbook.get_sheet_names()
    """

    def __init__(self, stream_or_string, **options):
        super(Deserializer, self).__init__(stream_or_string, **options)
        self.workbook = load_workbook(stream_or_string)
        self._select_first_sheet()

    def __next__(self):

        if not self._current_row_is_valid():
            """ 
            If the current row is not valid, next will try to go to the next sheet
            and deserialize the first object.
            """
            try:
                self._select_next_sheet()
            except:
                raise StopIteration

        # Build object:
        if self._current_row_is_valid():
            values = self._row_to_dict(self.workbook.active[self.current_row], self.fields)
            obj = base.build_instance(self.model_class, values, False)
            obj.save()
        else:
            # Sheets without data (only with header):
            raise StopIteration

        # Updata row position:
        self.current_row += 1

        return base.DeserializedObject(obj, {})

    def _has_next_sheet(self):
        sheets = self.workbook.get_sheet_names()
        return sheets.index(self.workbook.active.title) < (len(sheets)-1)

    def _select_sheet(self, sheet_name):
        self.workbook.active = self.workbook[sheet_name]
        self._start_sheet()

    def _select_first_sheet(self):
        self._select_sheet(self.workbook.sheetnames[0])

    def _select_next_sheet(self):
        if not self._has_next_sheet():
            raise Exception('Workbook can not change_to_next_sheet')
        current_index = self.workbook.sheetnames.index(self.workbook.active.title)
        next_sheet_name = self.workbook.sheetnames[current_index+1]
        self._select_sheet(next_sheet_name)

    def _start_sheet(self):
        """
        start all the properties associated with the active sheet, including:
        model_class, model_fields, num_fields, num_objects, fields, current_row.
        """
        self.model_class = self._get_model(self.workbook.active)
        self.model_fields = dict([(mf.name, mf) for mf in self.model_class._meta.fields])
        self.num_fields = len(self.workbook.active['1'])        # number of columns
        self.num_objects = len(self.workbook.active['A']) - 1   # number of rows - 1
        self.fields = [(cell.value, self.model_fields[cell.value]) for cell in self.workbook.active[1]]
        self._reset_current_row()


    def _current_row_is_valid(self):
        return self.current_row <= self.num_objects + 1

    def _reset_current_row(self):
        """
        first row with content is row number 2.
        Row index starts at 1, and first row is header.
        """
        self.current_row = 2

    def _get_model(self, sheet):
        """
        Each sheet on the workbook represents a model.
        Returns the model class from a specific sheet,
        or raises when the model has been not found.
        """

        if not sheet.title:
            raise base.DeserializationError("Worksheet is missing the required model name")
        try:
            return apps.get_model(sheet.title)
        except (LookupError, TypeError):
            raise base.DeserializationError('Worksheet has invalid model identifier: {}'.format(sheet.title))

    def _row_to_dict(self, row, fields):
        """
        Convert a spreadsheet row to a python dictionary taking into
        consideration the current model and field order (previous extracted from
        the spreadsheet's header).
        """
        values = {}

        # Getting all values from each column (on the row):
        for index, (field_name, field) in enumerate(fields):
            cell = row[index]
            values[field_name] = self._get_value(cell, field)

        return values

    def _get_value(self, cell, field):
        """
        Returns the value of the cell casted to the proper field type
        :param cell:
        :param field:
        :return:
        """

        # Empty cell:
        if cell.value is None:
            return None

        # Cell with excel formula:
        if str(cell.value).startswith('='):
            raise base.DeserializationError("Formulas are not supported at this time. Cell %s%s" % (cell.row, cell.column))

        # Process each field type:
        elif isinstance(field, models.ForeignKey):
            return field.related_model.objects.get(pk=cell.value)

        elif isinstance(field, models.AutoField):
            return cell.value

        elif isinstance(field, models.BooleanField):
            return cell.value if type(cell.value) is bool else bool(cell.value)

        elif isinstance(field, models.DateTimeField):
            # Order matters, DateTime needs to come before Date because datetime subclasses date
            # isinstance(aDateTimeInstance, DateField) is True, isinstance(aDateInstance, DateTimeField) is False
            if type(cell.value) is datetime:
                if settings.USE_TZ and timezone.is_naive(cell.value):
                    return timezone.make_aware(cell.value, pytz.UTC)
                else:
                    return cell.value
            else:
                # Handle a couple different timestamp formats -- first one is how *we* save 'em
                for pattern, format in DATETIME_FORMATS.items():
                    if re.match(pattern, str(cell.value)):
                        return datetime.strptime(str(cell.value), format)
                return str(cell.value)  # hope for the best

        elif isinstance(field, models.DateField):
            if type(cell.value) is datetime:
                return cell.value.date()
            else:
                return cell.value if type(cell.value) is date else datetime.strptime(cell.value, Y_M_D_FORMAT).date()

        elif isinstance(field, models.DecimalField):
            try:
                return decimal.Decimal(cell.value)
            except Exception as ex:
                print(cell.value)

        else:
            return str(cell.value)