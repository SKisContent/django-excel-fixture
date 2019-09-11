"""
Serializes a Django Model to Excel and vice-versa
Some observations:
OpenPyXL offers both reference addressing (e.g., 'A1', 'AC93') and indexed addressing.
When indexing into a row, the base (first) index is 1. However, after that you're
in the world of python tuples, so the indexes are 0-based (e.g., ws[1][0] == ws['A1'], ws[2][4] == ws['E2'])
The same way, columns can be directly referenced (ws['A']), but then rows are 0-based
(e.g., ws['A'][0] == ws['A1'], ws['C'][4] == ws['C5'])
"""
import decimal
import io
import re
from datetime import datetime, date

import pytz
from django.apps import apps
from django.conf import settings
from django.core.serializers import base
from django.utils import timezone
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.writer.excel import save_workbook

from django.db import models
from django.db.models import AutoField, BooleanField, CharField, DurationField, DateTimeField, ForeignKey, DecimalField, \
    IntegerField, PositiveIntegerField

PREFERRED_TS_FORMAT = '%Y-%m-%dT%H:%M:%S:%f%z'
Y_M_D_FORMAT = '%Y-%m-%d'

DATETIME_FORMATS = {r'\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}.\d{6}[-+]\d{4}':PREFERRED_TS_FORMAT,
                    r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}.\d{6}':'%Y-%m-%d %H:%M:%S:%f',
                    r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}':'%Y-%m-%d %H:%M:%S',
                    }

class ExcelMixin:
    workbook = 0
    cursor = 1
    headers = { }
    #  {
    #    '1ª Sheet Name': ['1ª column name', '2ª column name'],
    #    '2ª Sheet Name': ['1ª column name', '2ª column name'],
    #  }


    HEADER_ALIGN = Alignment(
        horizontal='center',
        vertical='center',
    )

    def column_index(self, sheet_title, name):
        return self.headers[sheet_title].index(name) + 1

    def increase_cursor(self):
        self.cursor += 1

    def create_workbook(self):

        print('create_workbook')

        self.wb = Workbook()
        self.ws = self.wb.active
        self.current_row = 2


    def select_sheet(self, name='Sheet 1'):

        print('select sheet', name)

        self.ws = self.wb[name]
        self.cursor = 1

    def current_sheet(self):
        return self.ws.title

    def has_sheet(self, name):
        if name  in self.wb.sheetnames:
            return True
        return False

    def create_sheet(self, name):

        print('\ncreate sheet:', name)

        if name not in self.headers:
            self.headers[name] = []

        self.wb.create_sheet(name)

    def create_hearder(self, list=[]):

        print('create header', list)

        for index, val in enumerate(list):
            self.ws.cell(row=1, column=(index+1), value=val.upper())
            self.headers[self.ws.title].append(val)

    def add_cell(self, obj, field, value):

        print('add cell', obj, ' - ', field, ' - ', value)
        print('data', self.ws.title,
              'name', field.name.upper(),
              'index', self.column_index(self.ws.title, field.name),
              )

        self.ws.cell(
            row=self.cursor,
            column=self.column_index(self.ws.title, field.name),
            value=value
        )

    def save_workbook(self, filename):

        print('save_workbook')
        print(self.headers)

        pass




class Serializer(ExcelMixin, base.Serializer):

    COMPATIBLE_FIELDS = [
        AutoField,
        BooleanField,
        CharField,
        DateTimeField,
        DecimalField,
        DurationField,
        ForeignKey,
        IntegerField,
        PositiveIntegerField,
    ]

    def _fields_dict_from_obj(self, obj):
        return dict([(mf.name, mf) for mf in obj._meta.fields if type(mf) in self.COMPATIBLE_FIELDS])

    def _fields_name_from_obj(self, obj):
        return [mf.name for mf in obj._meta.fields if type(mf) in self.COMPATIBLE_FIELDS]

    def start_serialization(self):
        self.create_workbook()

    def start_object(self, obj):

        if not hasattr(obj, "_meta"):
            raise base.SerializationError("Non-model object (%s) encountered during serialization" % type(obj))

        if not self.has_sheet(obj._meta.label):
            self.create_sheet(obj._meta.label)
            self.select_sheet(obj._meta.label)
            self.create_hearder(self._fields_name_from_obj(obj))

        self.increase_cursor()

    def handle_field(self, obj, field):

        try:
            if isinstance(field, models.BooleanField):
                value = field.value_from_object(obj)
            elif isinstance(field, models.DateTimeField):
                value = field.value_from_object(obj).strftime(PREFERRED_TS_FORMAT)
            elif isinstance(field, models.DateField):
                value = field.value_from_object(obj)
            elif isinstance(field, models.DecimalField):
                value = field.value_from_object(obj)
            else:
                value = field.value_to_string(obj)

            self.add_cell(obj, field, value)

        except base.SerializationError:
            raise ValueError("%s.%s (pk:%s) contains unserializable characters" %
                             (obj.__class__.__name__, field.name, obj.pk))

    def handle_fk_field(self, obj, field):

        #print('handle_fk_field: ', obj, ' - ', field)
        value = field.value_from_object(obj)
        self.add_cell(obj, field, value)

        pass

    def handle_m2m_field(self, obj, field):
        #print('handle_m2m_field: ', obj, ' - ', field)
        pass

    def end_serialization(self):

        print('end_serialization')

        self.save_workbook('')

        filename = self.stream.name
        if filename == '<stdout>':
            # If there is no file, dump a CSV representation to stdout
            out_str = self.csv(self.ws)
            self.stream.write(out_str)
        else:
            # The default stream is opened in text mode, but we need binary
            self.stream.close()
            # Just use the openpyxl saving method
            save_workbook(self.wb, filename)

    def csv(self, worksheet):

        print('cvs')

        """

        :param worksheet:
        :return:
        """
        num_rows = len(worksheet['A'])
        if num_rows == 1:
            raise base.SerializationError('There is no data to dump.')
        buffer = io.StringIO()
        for ix in range(num_rows):
            buffer.write(','.join([wrap(cell.value) if cell.value else '' for cell in worksheet[ix+1]]))
            buffer.write('\n')
        return buffer.getvalue()


def wrap(value):
    if value is None:
        return None
    if type(value) is str:
        value.replace('"','""')
    else:
        value = str(value)
    return '"{0}"'.format(value)


class Deserializer(base.Deserializer):
    def __init__(self, stream_or_string, **options):
        super(Deserializer, self).__init__(stream_or_string, **options)
        wb = load_workbook(stream_or_string)
        self.ws = wb.active
        self.model_identifier = self.ws['A1'].value
        self.Model = self._get_model(self.model_identifier)
        self.model_fields = dict([(mf.name, mf) for mf in self.Model._meta.fields])
        self.num_fields = len(self.ws['1'])
        self.num_objects = len(self.ws['A']) - 2
        self.fields = [(cell.value, self.model_fields[cell.value]) for cell in self.ws[2]]
        self.auto_now_fields = [af for af in self.Model._meta.fields
                                if (hasattr(af, 'auto_now') and af.auto_now) or
                                   (hasattr(af, 'auto_now_add') and af.auto_now_add)]

        self.current_row = 3 # First row is 1

    def __next__(self):
        if self.current_row < self.num_objects + 3:
            values = {}
            for ix in range(self.num_fields):
                values[self.fields[ix][0]] = self.get_value(self.ws[self.current_row][ix], self.fields[ix][1])
            present = datetime.now()
            # The following is not necessary since we are saving the object using it's save method
            # However, it doesn't hurt
            for af in self.auto_now_fields:
                if af.name not in values or values[af.name] is None:
                    values[af.name] = present
            obj = base.build_instance(self.Model, values, False)
            self.current_row += 1
            obj.save()
            return base.DeserializedObject(obj, {})
        raise StopIteration

    def _get_model(self, model_identifier):
        """
        Look up model
        """
        if not model_identifier:
            raise base.DeserializationError("Worksheet is missing the required model name")
        try:
            return apps.get_model(model_identifier)
        except (LookupError, TypeError):
            raise base.DeserializationError("Worksheet has invalid model identifier: '%s'"
                % (model_identifier))

    def get_value(self, cell, field):
        if cell.value is None:
            return None
        if str(cell.value).startswith('='):
            raise base.DeserializationError("Formulas are not supported at this time. Cell %s%s" % (cell.row, cell.column))
        if isinstance(field, models.BooleanField):
            return cell.value if type(cell.value) is bool else bool(cell.value)
        elif isinstance(field, models.DateTimeField):
            # Order matters, DateTime needs to come before Date because datetime subclasses date
            # isinstance(aDateTimeInstance, DateField) is True, isinstance(aDateInstance, DateTimeField) is False
            if type(cell.value) is datetime:
                if settings.USE_TZ and timezone.is_naive(cell.value):
                    return timezone.make_aware(cell.value, pytz.UTC)
                else:
                    return cell.value
            else:
                # Handle a couple different timestamp formats -- first one is how *we* save 'em
                for pattern, format in DATETIME_FORMATS.items():
                    if re.match(pattern, str(cell.value)):
                        return datetime.strptime(str(cell.value), format)
                return str(cell.value)  # hope for the best
        elif isinstance(field, models.DateField):
            if type(cell.value) is datetime:
                return cell.value.date()
            else:
                return cell.value if type(cell.value) is date else datetime.strptime(cell.value, Y_M_D_FORMAT).date()
        elif isinstance(field, models.DecimalField):
            try:
                return decimal.Decimal(cell.value)
            except Exception as ex:
                print(cell.value)
        else:
            return str(cell.value)
